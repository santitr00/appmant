"""
Exportación a Excel (.xlsx) y PDF — diario, semanal y mensual.

Rutas:
  GET /reportes/excel/dia?fecha=YYYY-MM-DD
  GET /reportes/excel/semana?fecha=YYYY-MM-DD   (lunes de la semana)
  GET /reportes/excel/mes?year=YYYY&month=MM
  GET /reportes/pdf/dia?fecha=YYYY-MM-DD
  GET /reportes/pdf/semana?fecha=YYYY-MM-DD
  GET /reportes/pdf/mes?year=YYYY&month=MM
"""
import calendar as cal_module
import io
from datetime import date, timedelta

from flask import make_response, render_template, request
from flask_login import current_user, login_required

from app.calendario.routes import DIAS_ES, MESES_ES, fecha_es
from app.models import PlantillaTarea, TareaExcepcional, TareaProgramada
from app.reportes import reportes_bp


# ── helpers ──────────────────────────────────────────────────────────────────

def _parse_fecha():
    try:
        return date.fromisoformat(request.args.get("fecha", date.today().isoformat()))
    except ValueError:
        return date.today()


def _parse_mes():
    hoy = date.today()
    try:
        year = int(request.args.get("year", hoy.year))
        month = int(request.args.get("month", hoy.month))
        if not (1 <= month <= 12):
            raise ValueError
        return year, month
    except (ValueError, TypeError):
        return hoy.year, hoy.month


def _tareas_del_dia(barrio_id, fecha):
    prog = (
        TareaProgramada.query
        .join(PlantillaTarea)
        .filter(PlantillaTarea.barrio_id == barrio_id, TareaProgramada.fecha == fecha)
        .order_by(TareaProgramada.horario)
        .all()
    )
    exc = (
        TareaExcepcional.query
        .filter_by(barrio_id=barrio_id, fecha=fecha)
        .order_by(TareaExcepcional.horario)
        .all()
    )
    return prog, exc


def _lunes_de(fecha: date) -> date:
    return fecha - timedelta(days=fecha.weekday())


def _semana_dias(fecha: date):
    lunes = _lunes_de(fecha)
    return [lunes + timedelta(days=i) for i in range(7)]


# ── Excel diario ─────────────────────────────────────────────────────────────

@reportes_bp.route("/excel/dia")
@login_required
def excel_dia():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fecha = _parse_fecha()
    prog, exc = _tareas_del_dia(current_user.barrio_id, fecha)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tareas del día"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Reporte diario — {fecha_es(fecha)} — {current_user.barrio.nombre}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Horario", "Tarea / Descripción", "Tipo", "Responsable", "Estado", "Justificación"]
    hfill = PatternFill("solid", fgColor="2563EB")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    row = 4
    for t in prog:
        just = t.justificacion.motivo if t.justificacion else ""
        ws.cell(row=row, column=1, value=t.horario.strftime("%H:%M"))
        ws.cell(row=row, column=2, value=t.plantilla.nombre)
        ws.cell(row=row, column=3, value="Rutinaria")
        ws.cell(row=row, column=4, value=t.responsable or "")
        ws.cell(row=row, column=5, value=t.estado.replace("_", " "))
        ws.cell(row=row, column=6, value=just)
        row += 1

    for t in exc:
        ws.cell(row=row, column=1, value=t.horario.strftime("%H:%M"))
        ws.cell(row=row, column=2, value=t.descripcion)
        ws.cell(row=row, column=3, value=f"Excepcional ({t.prioridad})")
        ws.cell(row=row, column=4, value=t.responsable or "")
        ws.cell(row=row, column=5, value=t.estado)
        ws.cell(row=row, column=6, value="")
        row += 1

    ws.cell(row=row + 1, column=1, value="Total tareas:").font = Font(bold=True)
    ws.cell(row=row + 1, column=2, value=len(prog) + len(exc))

    for i, w in enumerate([10, 42, 18, 22, 16, 42], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = make_response(buf.read())
    resp.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="reporte_{fecha.isoformat()}.xlsx"'
    )
    return resp


# ── Excel semanal ─────────────────────────────────────────────────────────────

@reportes_bp.route("/excel/semana")
@login_required
def excel_semana():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fecha = _parse_fecha()
    dias = _semana_dias(fecha)
    lunes = dias[0]
    domingo = dias[6]
    barrio_id = current_user.barrio_id

    wb = Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen semanal"
    hfill = PatternFill("solid", fgColor="2563EB")

    titulo = (
        f"Reporte semanal — {lunes.strftime('%d/%m/%Y')} al {domingo.strftime('%d/%m/%Y')}"
        f" — {current_user.barrio.nombre}"
    )
    ws.merge_cells("A1:G1")
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Fecha", "Día", "Completadas", "No realizadas", "Pendientes", "Total prog.", "Excepcionales"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    totales = [0, 0, 0, 0, 0]
    row = 4
    for d in dias:
        prog, exc = _tareas_del_dia(barrio_id, d)
        completadas = sum(1 for t in prog if t.estado == "completada")
        no_real = sum(1 for t in prog if t.estado == "no_realizada")
        pendientes = sum(1 for t in prog if t.estado == "pendiente")

        ws.cell(row=row, column=1, value=d.strftime("%d/%m/%Y"))
        ws.cell(row=row, column=2, value=DIAS_ES[d.weekday()])
        ws.cell(row=row, column=3, value=completadas)
        ws.cell(row=row, column=4, value=no_real)
        ws.cell(row=row, column=5, value=pendientes)
        ws.cell(row=row, column=6, value=len(prog))
        ws.cell(row=row, column=7, value=len(exc))
        totales[0] += completadas
        totales[1] += no_real
        totales[2] += pendientes
        totales[3] += len(prog)
        totales[4] += len(exc)
        row += 1

    ws.cell(row=row + 1, column=1, value="TOTALES").font = Font(bold=True)
    for i, v in enumerate(totales, 3):
        ws.cell(row=row + 1, column=i, value=v).font = Font(bold=True)

    for i, w in enumerate([14, 12, 14, 16, 12, 13, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 2: Detalle de tareas ────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalle tareas")
    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"Detalle — {lunes.strftime('%d/%m/%Y')} al {domingo.strftime('%d/%m/%Y')}"
    ws2["A1"].font = Font(bold=True, size=12)

    for col, h in enumerate(["Fecha", "Día", "Tarea / Descripción", "Tipo", "Responsable", "Estado"], 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill

    r = 4
    for d in dias:
        prog, exc = _tareas_del_dia(barrio_id, d)
        for t in prog:
            ws2.cell(row=r, column=1, value=d.strftime("%d/%m/%Y"))
            ws2.cell(row=r, column=2, value=DIAS_ES[d.weekday()])
            ws2.cell(row=r, column=3, value=t.plantilla.nombre)
            ws2.cell(row=r, column=4, value="Rutinaria")
            ws2.cell(row=r, column=5, value=t.responsable or "")
            ws2.cell(row=r, column=6, value=t.estado.replace("_", " "))
            r += 1
        for t in exc:
            ws2.cell(row=r, column=1, value=d.strftime("%d/%m/%Y"))
            ws2.cell(row=r, column=2, value=DIAS_ES[d.weekday()])
            ws2.cell(row=r, column=3, value=t.descripcion)
            ws2.cell(row=r, column=4, value=f"Excepcional ({t.prioridad})")
            ws2.cell(row=r, column=5, value=t.responsable or "")
            ws2.cell(row=r, column=6, value=t.estado)
            r += 1

    for i, w in enumerate([14, 12, 40, 18, 22, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = make_response(buf.read())
    resp.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="semana_{lunes.isoformat()}.xlsx"'
    )
    return resp


# ── Excel mensual ─────────────────────────────────────────────────────────────

@reportes_bp.route("/excel/mes")
@login_required
def excel_mes():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    year, month = _parse_mes()
    _, last_day = cal_module.monthrange(year, month)
    barrio_id = current_user.barrio_id

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen mensual"
    hfill = PatternFill("solid", fgColor="2563EB")

    nombre_mes = MESES_ES[month].capitalize()
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Reporte mensual — {nombre_mes} {year} — {current_user.barrio.nombre}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Fecha", "Día", "Completadas", "No realizadas", "Pendientes", "Total prog.", "Excepcionales"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    justificaciones = []
    row = 4
    totales = [0, 0, 0, 0, 0]

    for day in range(1, last_day + 1):
        fecha = date(year, month, day)
        prog, exc = _tareas_del_dia(barrio_id, fecha)
        completadas = sum(1 for t in prog if t.estado == "completada")
        no_real = sum(1 for t in prog if t.estado == "no_realizada")
        pendientes = sum(1 for t in prog if t.estado == "pendiente")

        ws.cell(row=row, column=1, value=fecha.strftime("%d/%m/%Y"))
        ws.cell(row=row, column=2, value=DIAS_ES[fecha.weekday()])
        ws.cell(row=row, column=3, value=completadas)
        ws.cell(row=row, column=4, value=no_real)
        ws.cell(row=row, column=5, value=pendientes)
        ws.cell(row=row, column=6, value=len(prog))
        ws.cell(row=row, column=7, value=len(exc))

        totales[0] += completadas
        totales[1] += no_real
        totales[2] += pendientes
        totales[3] += len(prog)
        totales[4] += len(exc)

        for t in prog:
            if t.justificacion:
                justificaciones.append({
                    "fecha": fecha.strftime("%d/%m/%Y"),
                    "tarea": t.plantilla.nombre,
                    "motivo": t.justificacion.motivo,
                    "exc": t.justificacion.tarea_excepcional.descripcion[:50]
                    if t.justificacion.tarea_excepcional else "",
                })
        row += 1

    ws.cell(row=row + 1, column=1, value="TOTALES").font = Font(bold=True)
    for i, v in enumerate(totales, 3):
        ws.cell(row=row + 1, column=i, value=v).font = Font(bold=True)

    for i, w in enumerate([14, 12, 14, 16, 12, 13, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Hoja 2: detalle de tareas
    ws2 = wb.create_sheet("Detalle tareas")
    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"Detalle de tareas — {nombre_mes} {year}"
    ws2["A1"].font = Font(bold=True, size=12)
    for col, h in enumerate(["Fecha", "Día", "Tarea / Descripción", "Tipo", "Responsable", "Estado"], 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill

    r = 4
    for day in range(1, last_day + 1):
        fecha = date(year, month, day)
        prog, exc = _tareas_del_dia(barrio_id, fecha)
        for t in prog:
            ws2.cell(row=r, column=1, value=fecha.strftime("%d/%m/%Y"))
            ws2.cell(row=r, column=2, value=DIAS_ES[fecha.weekday()])
            ws2.cell(row=r, column=3, value=t.plantilla.nombre)
            ws2.cell(row=r, column=4, value="Rutinaria")
            ws2.cell(row=r, column=5, value=t.responsable or "")
            ws2.cell(row=r, column=6, value=t.estado.replace("_", " "))
            r += 1
        for t in exc:
            ws2.cell(row=r, column=1, value=fecha.strftime("%d/%m/%Y"))
            ws2.cell(row=r, column=2, value=DIAS_ES[fecha.weekday()])
            ws2.cell(row=r, column=3, value=t.descripcion)
            ws2.cell(row=r, column=4, value=f"Excepcional ({t.prioridad})")
            ws2.cell(row=r, column=5, value=t.responsable or "")
            ws2.cell(row=r, column=6, value=t.estado)
            r += 1

    for i, w in enumerate([14, 12, 40, 18, 22, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Hoja 3: justificaciones
    ws3 = wb.create_sheet("Justificaciones")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = f"Justificaciones — {nombre_mes} {year}"
    ws3["A1"].font = Font(bold=True, size=12)
    for col, h in enumerate(["Fecha", "Tarea", "Motivo", "Vinculada a excepcional"], 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
    for r, j in enumerate(justificaciones, 4):
        ws3.cell(row=r, column=1, value=j["fecha"])
        ws3.cell(row=r, column=2, value=j["tarea"])
        ws3.cell(row=r, column=3, value=j["motivo"])
        ws3.cell(row=r, column=4, value=j["exc"])
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 35
    ws3.column_dimensions["C"].width = 50
    ws3.column_dimensions["D"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = make_response(buf.read())
    resp.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="reporte_{year}_{month:02d}.xlsx"'
    )
    return resp


# ── PDF diario ────────────────────────────────────────────────────────────────

@reportes_bp.route("/pdf/dia")
@login_required
def pdf_dia():
    from xhtml2pdf import pisa

    fecha = _parse_fecha()
    prog, exc = _tareas_del_dia(current_user.barrio_id, fecha)

    html = render_template(
        "reportes/pdf_dia.html",
        fecha=fecha,
        fecha_label=fecha_es(fecha),
        tareas_prog=prog,
        tareas_exc=exc,
        barrio=current_user.barrio,
    )
    buf = io.BytesIO()
    pisa.CreatePDF(html, dest=buf)
    buf.seek(0)

    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="reporte_{fecha.isoformat()}.pdf"'
    )
    return resp


# ── PDF semanal ───────────────────────────────────────────────────────────────

@reportes_bp.route("/pdf/semana")
@login_required
def pdf_semana():
    from xhtml2pdf import pisa

    fecha = _parse_fecha()
    dias_list = _semana_dias(fecha)
    lunes = dias_list[0]
    domingo = dias_list[6]
    barrio_id = current_user.barrio_id

    dias = []
    for d in dias_list:
        prog, exc = _tareas_del_dia(barrio_id, d)
        dias.append({
            "fecha": d,
            "fecha_str": d.strftime("%d/%m/%Y"),
            "dia_nombre": DIAS_ES[d.weekday()],
            "completadas": sum(1 for t in prog if t.estado == "completada"),
            "no_realizadas": sum(1 for t in prog if t.estado == "no_realizada"),
            "pendientes": sum(1 for t in prog if t.estado == "pendiente"),
            "total_prog": len(prog),
            "excepcionales": len(exc),
            "tareas_prog": prog,
            "tareas_exc": exc,
        })

    html = render_template(
        "reportes/pdf_semana.html",
        dias=dias,
        lunes=lunes,
        domingo=domingo,
        barrio=current_user.barrio,
    )
    buf = io.BytesIO()
    pisa.CreatePDF(html, dest=buf)
    buf.seek(0)

    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="semana_{lunes.isoformat()}.pdf"'
    )
    return resp


# ── PDF mensual ───────────────────────────────────────────────────────────────

@reportes_bp.route("/pdf/mes")
@login_required
def pdf_mes():
    from xhtml2pdf import pisa

    year, month = _parse_mes()
    _, last_day = cal_module.monthrange(year, month)
    barrio_id = current_user.barrio_id
    nombre_mes = MESES_ES[month].capitalize()

    dias = []
    justificaciones = []

    for day in range(1, last_day + 1):
        fecha = date(year, month, day)
        prog, exc = _tareas_del_dia(barrio_id, fecha)
        dias.append({
            "fecha": fecha,
            "fecha_str": fecha.strftime("%d/%m/%Y"),
            "dia_nombre": DIAS_ES[fecha.weekday()],
            "completadas": sum(1 for t in prog if t.estado == "completada"),
            "no_realizadas": sum(1 for t in prog if t.estado == "no_realizada"),
            "pendientes": sum(1 for t in prog if t.estado == "pendiente"),
            "total_prog": len(prog),
            "excepcionales": len(exc),
            "tareas_prog": prog,
            "tareas_exc": exc,
        })
        for t in prog:
            if t.justificacion:
                justificaciones.append({
                    "fecha_str": fecha.strftime("%d/%m/%Y"),
                    "tarea": t.plantilla.nombre,
                    "motivo": t.justificacion.motivo,
                    "exc": t.justificacion.tarea_excepcional.descripcion[:60]
                    if t.justificacion.tarea_excepcional else "",
                })

    html = render_template(
        "reportes/pdf_mes.html",
        year=year,
        month=month,
        nombre_mes=nombre_mes,
        dias=dias,
        justificaciones=justificaciones,
        barrio=current_user.barrio,
    )
    buf = io.BytesIO()
    pisa.CreatePDF(html, dest=buf)
    buf.seek(0)

    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="reporte_{year}_{month:02d}.pdf"'
    )
    return resp
